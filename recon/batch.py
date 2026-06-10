"""Batch comparison driven by a manifest: rows of (index, source path, target path)."""

import csv
from dataclasses import dataclass
from pathlib import Path

from .core import CompareError, compare_paths
from .loaders import LoadError
from .recondiff import ReconError, ReconResult
from .textdiff import DiffResult


class ManifestError(Exception):
    pass


@dataclass
class BatchItem:
    index: int
    source: str
    target: str
    result: DiffResult | ReconResult | None = None
    error: str | None = None

    @property
    def status(self) -> str:
        if self.error is not None:
            return "ERROR"
        if isinstance(self.result, ReconResult):
            return "MATCH" if self.result.reconciled else "DIFF"
        return "MATCH" if self.result.identical else "DIFF"

    @property
    def detail(self) -> str:
        if self.error is not None:
            return self.error
        r = self.result
        if isinstance(r, ReconResult):
            if r.reconciled:
                return f"{r.matched} records reconciled"
            parts = []
            if r.mismatched_keys:
                parts.append(f"{len(r.mismatched_keys)} mismatched")
            if r.only_in_a:
                parts.append(f"{len(r.only_in_a)} only in source")
            if r.only_in_b:
                parts.append(f"{len(r.only_in_b)} only in target")
            if r.duplicate_keys_a or r.duplicate_keys_b:
                parts.append(f"{len(r.duplicate_keys_a) + len(r.duplicate_keys_b)} duplicate keys")
            return ", ".join(parts)
        if r.identical:
            return "identical"
        return f"+{r.added} -{r.removed} ~{r.changed} lines"


MANIFEST_HEADERS = ["index", "source", "target"]


def _read_xlsx_rows(path: Path) -> list[list[str]]:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ManifestError(f"{path}: cannot read .xlsx: {e}") from e
    rows = [["" if c is None else str(c) for c in row]
            for row in wb.active.iter_rows(values_only=True)]
    wb.close()
    return rows


def load_manifest(path: str | Path) -> list[tuple[str, str]]:
    """Read (source, target) pairs from a .xlsx or .csv manifest.

    Column layout: a 'source' and 'target' column found by header name, or
    positionally — columns 2 and 3 when an index column is present, else 1 and 2.
    The index column is informational only; rows are renumbered 1..N in order.
    """
    path = Path(path)
    if not path.is_file():
        raise ManifestError(f"{path}: manifest not found")
    if path.suffix.lower() == ".xlsx":
        rows = _read_xlsx_rows(path)
    elif path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            rows = [[c or "" for c in row] for row in csv.reader(f)]
    else:
        raise ManifestError(f"{path}: manifest must be .xlsx or .csv")

    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        raise ManifestError(f"{path}: manifest is empty")

    header = [str(c).strip().lower() for c in rows[0]]
    src_col = tgt_col = None
    for i, name in enumerate(header):
        if src_col is None and "source" in name:
            src_col = i
        elif tgt_col is None and "target" in name:
            tgt_col = i
    if src_col is not None and tgt_col is not None:
        data = rows[1:]
    else:
        # no recognizable header: positional, all rows are data
        width = max(len(r) for r in rows)
        src_col, tgt_col = (1, 2) if width >= 3 else (0, 1)
        data = rows

    pairs = []
    for row in data:
        src = str(row[src_col]).strip() if len(row) > src_col else ""
        tgt = str(row[tgt_col]).strip() if len(row) > tgt_col else ""
        if not src and not tgt:
            continue
        if not src or not tgt:
            raise ManifestError(
                f"{path}: row {len(pairs) + 1} needs both source and target "
                f"(got source={src!r}, target={tgt!r})"
            )
        pairs.append((src, tgt))
    if not pairs:
        raise ManifestError(f"{path}: no file pairs found")
    return pairs


def run_batch(
    pairs: list[tuple[str, str]],
    key: str | None = None,
    ignore: str | None = None,
    tolerance: float = 0.0,
    base_dir: str | Path | None = None,
) -> list[BatchItem]:
    """Compare each (source, target) pair; relative paths resolve against base_dir.

    A global --key only applies to tabular pairs; non-tabular pairs in the same
    batch fall back to a text diff instead of failing.
    """
    base = Path(base_dir) if base_dir else Path.cwd()
    items = []
    for n, (src, tgt) in enumerate(pairs, start=1):
        item = BatchItem(index=n, source=src, target=tgt)
        path_a, path_b = Path(src).expanduser(), Path(tgt).expanduser()
        if not path_a.is_absolute():
            path_a = base / path_a
        if not path_b.is_absolute():
            path_b = base / path_b
        try:
            try:
                item.result = compare_paths(path_a, path_b, key=key,
                                            ignore=ignore, tolerance=tolerance)
            except CompareError:
                # key given but this pair isn't tabular: plain text diff
                item.result = compare_paths(path_a, path_b)
        except (LoadError, ReconError, OSError) as e:
            item.error = str(e)
        items.append(item)
    return items


def write_template_xlsx(path: str | Path) -> None:
    """Write a starter manifest as an Excel table; the index column auto-fills
    when new rows are added inside the table."""
    import openpyxl
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "manifest"
    ws.append(["index", "source", "target"])
    for row in range(2, 6):
        ws.cell(row=row, column=1, value="=ROW()-1")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    table = Table(displayName="manifest", ref="A1:C5")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    wb.save(path)
