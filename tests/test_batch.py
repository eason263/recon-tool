from pathlib import Path

import pytest

from recon.batch import ManifestError, load_manifest, run_batch, write_template_xlsx

FIXTURES = Path(__file__).parent / "fixtures"


def test_load_manifest_csv_with_header(tmp_path):
    m = tmp_path / "m.csv"
    m.write_text("index,source,target\n1,a.csv,b.csv\n2,a.txt,b.txt\n")
    assert load_manifest(m) == [("a.csv", "b.csv"), ("a.txt", "b.txt")]


def test_load_manifest_csv_positional_no_header(tmp_path):
    m = tmp_path / "m.csv"
    m.write_text("a.csv,b.csv\na.txt,b.txt\n")
    assert load_manifest(m) == [("a.csv", "b.csv"), ("a.txt", "b.txt")]


def test_load_manifest_xlsx(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["index", "source", "target"])
    ws.append([1, "a.json", "b.json"])
    ws.append([None, None, None])  # blank row ignored
    ws.append([2, "a.xml", "b.xml"])
    m = tmp_path / "m.xlsx"
    wb.save(m)
    assert load_manifest(m) == [("a.json", "b.json"), ("a.xml", "b.xml")]


def test_load_manifest_missing_target_raises(tmp_path):
    m = tmp_path / "m.csv"
    m.write_text("index,source,target\n1,a.csv,\n")
    with pytest.raises(ManifestError):
        load_manifest(m)


def test_run_batch_mixed_types_and_errors():
    pairs = [
        ("a.csv", "b.csv"),    # tabular: reconciled with key
        ("a.txt", "b.txt"),    # text: key falls back to diff
        ("a.txt", "a.txt"),    # identical
        ("missing.txt", "b.txt"),  # error
    ]
    items = run_batch(pairs, key="trade_id", ignore="updated_at", base_dir=FIXTURES)
    assert [i.status for i in items] == ["DIFF", "DIFF", "MATCH", "ERROR"]
    assert [i.index for i in items] == [1, 2, 3, 4]
    assert "only in source" in items[0].detail
    assert "lines" in items[1].detail
    assert items[3].error is not None


def test_template_roundtrip(tmp_path):
    out = tmp_path / "manifest.xlsx"
    write_template_xlsx(out)
    import openpyxl

    ws = openpyxl.load_workbook(out).active
    assert [c.value for c in ws[1]] == ["index", "source", "target"]
    assert ws["A2"].value == "=ROW()-1"
